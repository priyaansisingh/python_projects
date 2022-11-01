import streamlit as st 
import os
os.system('cls')
import csv
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, alignment, borders
from openpyxl.styles import Alignment 
from openpyxl.cell import Cell
from openpyxl.styles.colors import Color, RgbColor
import email, smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders 

def file_selector_master(folder_path='./sample_input/'):
    filenames_master = os.listdir(folder_path)
    selected_filename = st.selectbox('Browse for master_roll.csv', filenames_master)
    return os.path.join(folder_path, selected_filename)

def file_selector_response(folder_path='./sample_input/'):
    filenames_response = os.listdir(folder_path)
    selected_filename = st.selectbox('Browse for response.csv', filenames_response)
    return os.path.join(folder_path, selected_filename)

#marksheet generator 

def sheets(response_list, ANSWER, positive, negative, absent): #
    make_excel_present(response_list, ANSWER, positive, negative)
    make_excel_present(absent, ANSWER, positive, negative)
    answer(ANSWER, positive, negative)

def make_excel_present(response_list, ANSWER, positive, negative):
    for items in response_list:
        wb=Workbook()
        sheet=wb.active

        #layout
        layout=(
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['Name:', items[3], '', 'Exam:', 'quiz'],
            ['Roll No.:', items[6], '', '', '',],
            ['', '', '', '', '',],
            ['', 'Right', 'Wrong', 'Not Attempt', 'Max',],
            ['No.', items[-1][0], items[-1][1], items[-1][2], items[-1][5],],
            ['Marking', positive, negative, 0,],
            ['Total', items[-1][3], items[-1][4], '', items[-1][6]],
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['Student Ans', 'Correct Ans', '', '', '',],
        )

        for row in layout:
            sheet.append(row)

        sheet['C10'].font=Font(color=Color('FF0000'))
        sheet['C11'].font=Font(color=Color('FF0000'))
        sheet['C11'].font=Font(color=Color('FF0000'))
        sheet['C12'].font=Font(color=Color('FF0000'))
        sheet['B10'].font=Font(color=Color('008000'))
        sheet['B11'].font=Font(color=Color('008000'))
        sheet['B12'].font=Font(color=Color('008000'))
        sheet['E12'].font=Font(color=Color('0000FF'))

        sheet.merge_cells('B6:C6')

        #Bold and Underline
        sheet['B6'].font = Font(bold=True)
        sheet['B7'].font = Font(bold=True)
        sheet['B9'].font = Font(bold=True)
        sheet['A10'].font = Font(bold=True)
        sheet['A11'].font = Font(bold=True)
        sheet['A12'].font = Font(bold=True)
        sheet['A15'].font = Font(bold=True)
        sheet['B15'].font = Font(bold=True)
        sheet['C9'].font = Font(bold=True)
        sheet['D9'].font = Font(bold=True)
        sheet['E9'].font = Font(bold=True)
        sheet['E6'].font = Font(bold=True)
        sheet['A5'].font = Font(bold=True, underline='single')

        #Merging Cells
        sheet.merge_cells('A5:E5')  
        cell = sheet.cell(row=5, column=1) 
        fontStyle = Font(size = "18", underline='single', bold=True) 
        sheet.cell(row = 5, column = 1).font = fontStyle 
        cell.value = 'Mark sheet' 
        
        cell.alignment = Alignment(horizontal='center', vertical='bottom')

        for j in range(len(ANSWER)-7):
            if items[j+7]!=ANSWER[j+7]:                
                color1='FF0000'
            else:
                color1='008000'
            sheet.cell(row=j+16, column=1, value=items[j+7]).font=Font(color=Color(color1))
            
            sheet.cell(row=j+16, column=2, value=ANSWER[j+7]).font=Font(color=Color('0000FF'))
            
            
        #image
        img = openpyxl.drawing.image.Image('iitp_logo_heading.jpg')
        img.anchor = 'A1'
        sheet.add_image(img)
        wb.save("C:\\Users\\Dell\\Documents\\GitHub\\temp\\1901CB33_2021_(1)\\proj1\\sample_output\\marksheet\\" + items[6] + ".xlsx")

def make_excel_absent(absent, ANSWER, positive, negative):
    for items in absent:
        wb=Workbook()
        sheet=wb.active

        #layout
        layout=(
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['Name:', items[1], '', 'Exam:', 'quiz'],
            ['Roll No.:', items[0], '', '', '',],
            ['', '', '', '', '',],
            ['', 'Right', 'Wrong', 'Not Attempt', 'Max',],
            ['No.', '', '', '', '',],
            ['Marking', positive, negative, 0,],
            ['Total', '', '', '', ''],
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['Student Ans', 'Correct Ans', '', '', '',],
        )

        for row in layout:
            sheet.append(row)

        sheet['C10'].font=Font(color=Color('FF0000'))
        sheet['C11'].font=Font(color=Color('FF0000'))
        sheet['C11'].font=Font(color=Color('FF0000'))
        sheet['C12'].font=Font(color=Color('FF0000'))
        sheet['B10'].font=Font(color=Color('008000'))
        sheet['B11'].font=Font(color=Color('008000'))
        sheet['B12'].font=Font(color=Color('008000'))
        sheet['E12'].font=Font(color=Color('0000FF'))
        
        sheet.merge_cells('B6:C6')

        #Bold and Underline
        sheet['B6'].font = Font(bold=True)
        sheet['B7'].font = Font(bold=True)
        sheet['B9'].font = Font(bold=True)
        sheet['A10'].font = Font(bold=True)
        sheet['A11'].font = Font(bold=True)
        sheet['A12'].font = Font(bold=True)
        sheet['A15'].font = Font(bold=True)
        sheet['B15'].font = Font(bold=True)
        sheet['C9'].font = Font(bold=True)
        sheet['D9'].font = Font(bold=True)
        sheet['E9'].font = Font(bold=True)
        sheet['E6'].font = Font(bold=True)
        sheet['A5'].font = Font(bold=True, underline='single')

        #Merging Cells
        sheet.merge_cells('A5:E5')  
        cell = sheet.cell(row=5, column=1) 
        fontStyle = Font(size = "18", underline='single', bold=True) 
        sheet.cell(row = 5, column = 1).font = fontStyle 
        cell.value = 'Mark sheet' 
        
        cell.alignment = Alignment(horizontal='center', vertical='bottom')

        for j in range(len(ANSWER)-7):
            sheet.cell(row=j+16, column=2, value=ANSWER[j+7]).font=Font(color=Color('0000FF'))
            
            
        #image
        img = openpyxl.drawing.image.Image('iitp_logo_heading.jpg')
        img.anchor = 'A1'
        sheet.add_image(img)
        wb.save("C:\\Users\\Dell\\Documents\\GitHub\\temp\\1901CB33_2021_(1)\\proj1\\sample_output\\marksheet\\" + items[0] + ".xlsx")

def concise(response_list, absent): 
    header=['Timestamp',	'Email address',	'Google_Score',	'Name',	'IITP webmail',	'Phone (10 digit only)',	'Score_After_Negative', 	'Roll Number',	'Unnamed: 7',	'Unnamed: 8',	'Unnamed: 9',	'Unnamed: 10',	'Unnamed: 11',	'Unnamed: 12',	'Unnamed: 13',	'Unnamed: 14',	'Unnamed: 15',	'Unnamed: 16',	'Unnamed: 17',	'Unnamed: 18',	'Unnamed: 19',	'Unnamed: 20',	'Unnamed: 21',	'Unnamed: 22',	'Unnamed: 23',	'Unnamed: 24',	'Unnamed: 25',	'Unnamed: 26',	'Unnamed: 27',	'Unnamed: 28',	'Unnamed: 29',	'Unnamed: 30',	'Unnamed: 31',	'Unnamed: 32',	'Unnamed: 33',	'Positive Marks',	'Total MArks']

    with open('C:\\Users\\Dell\\Documents\\GitHub\\temp\\1901CB33_2021_(1)\\proj1\\sample_output\\marksheet\\concise_marksheet.csv', "w", newline="") as f:
        writer=csv.writer(f)
        writer.writerow(header)
        for row in response_list:
            marks=row[-1][3]+row[-1][4]
            positive_marks=row[-1][3]
            row.pop()
            row.extend([positive_marks, marks])
            writer.writerow(row)

def answer(ANSWER, positive, negative):
    wb=Workbook()
    sheet=wb.active

    #layout
    layout=(
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['Name:', ANSWER[3], '', 'Exam:', 'quiz'],
            ['Roll No.:', ANSWER[6], '', '', '',],
            ['', '', '', '', '',],
            ['', 'Right', 'Wrong', 'Not Attempt', 'Max',],
            ['No.', len(ANSWER)-7, 0, 0, len(ANSWER)-7,],
            ['Marking', positive, negative, 0,],
            ['Total', positive*(len(ANSWER)-7), 0, '', str(positive*(len(ANSWER)-7))+'/'+str(positive*(len(ANSWER)-7))],
            ['', '', '', '', '',],
            ['', '', '', '', '',],
            ['Student Ans', 'Correct Ans', '', '', '',],
        )

    for row in layout:
        sheet.append(row)

    sheet['C10'].font=Font(color=Color('FF0000'))
    sheet['C11'].font=Font(color=Color('FF0000'))
    sheet['C11'].font=Font(color=Color('FF0000'))
    sheet['C12'].font=Font(color=Color('FF0000'))
    sheet['B10'].font=Font(color=Color('008000'))
    sheet['B11'].font=Font(color=Color('008000'))
    sheet['B12'].font=Font(color=Color('008000'))
    sheet['E12'].font=Font(color=Color('0000FF'))

    sheet.merge_cells('B6:C6')

    #Bold and Underline
    sheet['B6'].font = Font(bold=True)
    sheet['B7'].font = Font(bold=True)
    sheet['B9'].font = Font(bold=True)
    sheet['A10'].font = Font(bold=True)
    sheet['A11'].font = Font(bold=True)
    sheet['A12'].font = Font(bold=True)
    sheet['A15'].font = Font(bold=True)
    sheet['B15'].font = Font(bold=True)
    sheet['C9'].font = Font(bold=True)
    sheet['D9'].font = Font(bold=True)
    sheet['E9'].font = Font(bold=True)
    sheet['E6'].font = Font(bold=True)
    sheet['A5'].font = Font(bold=True, underline='single')

    #Merging Cells
    sheet.merge_cells('A5:E5')  
    cell = sheet.cell(row=5, column=1) 
    fontStyle = Font(size = "18", underline='single', bold=True) 
    sheet.cell(row = 5, column = 1).font = fontStyle 
    cell.value = 'Mark Sheet' 
    
    cell.alignment = Alignment(horizontal='center', vertical='bottom')

    for j in range(len(ANSWER)-7):
        sheet.cell(row=j+16, column=1, value=ANSWER[j+7]).font=Font(color=Color('008000'))
        sheet.cell(row=j+16, column=2, value=ANSWER[j+7]).font=Font(color=Color('0000FF'))
            
            
    #image
    img = openpyxl.drawing.image.Image('iitp_logo_heading.jpg')
    img.anchor = 'A1'
    sheet.add_image(img)
    wb.save("C:\\Users\\Dell\\Documents\\GitHub\\temp\\1901CB33_2021_(1)\\proj1\\sample_output\\marksheet\\" + 'ANSWER' + ".xlsx")   

def send_email(response_list):
    
    #The mail addresses and  password
    sender_address ='forprojectmailing@gmail.com'              
    sender_pass ='projectmailing'                                       

    for people in response_list:
        
        reciver_mail = [people[1], people[4]]
        
        print(reciver_mail)
        
        message = MIMEMultipart()
        message['From'] = sender_address
        message['To'] =  reciver_mail[0]          
        message['Subject'] =  'Your CS384 Marksheet'      
        

        mail_content = '''Hello Student,
        Please find your attached marksheet.'''

        message.attach(MIMEText(mail_content, 'plain')) 
        filename = "sample_output\\marksheet\\"+people[6]+".xlsx"  # In same directory as script                                                   #str(input('Enter File Name With Extension To Attchment :- '))

        with open(filename, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read()) 
        encoders.encode_base64(part) 

        part.add_header('Content-Disposition', "attachment; filename= %s" % filename) 

        message.attach(part) 

        s = smtplib.SMTP('smtp.gmail.com', 587) 
        s.starttls() 
        s.login(sender_address, sender_pass) 
        text = message.as_string()
        s.sendmail(sender_address, reciver_mail[0], text) 

        message['To'] = reciver_mail[1]
        s.sendmail(sender_address, reciver_mail[1], text)

        s.quit()

st.title('Marksheet Generator')

st.write("""
    Details
    """)

positive = st.number_input('Enter marks for correct answer',0.000)
negative = st.number_input('Enter -ve marks for wrong answer', min_value=None, max_value=0.000)

st.text("")
st.text("")
st.text("")

st.write("""
    Browse For Input
    """)

csv_file_1=st.file_uploader('Upload a .csv file (master_roll.csv)', type="csv")
if csv_file_1 is not None:
    # data_1=pd.read_csv(csv_file_1)     
    # df_1=pd.DataFrame(data_1)
    with open(os.path.join("sample_input",csv_file_1.name),"wb") as f: 
      f.write(csv_file_1.getbuffer())         
    st.success("Saved File")

csv_file_2=st.file_uploader('Upload a .csv file (responses.csv)', type="csv")
if (csv_file_2) is not None:      
    i=0
    # data_2=pd.read_csv(csv_file_2)     
    # df_2=pd.DataFrame(data_2)
    with open(os.path.join("sample_input",csv_file_2.name),"wb") as f: 
      f.write(csv_file_2.getbuffer())         
    st.success("Saved File")
    response_list=[]
    # for i in df_2.values: ##################
    #     j=i.tolist()
    #     response_list.append(j)
    with open(os.path.join("sample_input",csv_file_2.name),"r") as f: 
        reader=csv.reader(f)
        response_list=list(reader)
        for items in response_list:
            if items[6]=='ANSWER':
                i=1
                response_list.remove(items)
                ANSWER=items
        response_list.pop(0)
        max=len(ANSWER)-7

    if i==0:
        st.write("No Roll Number with ANSWER is present, Cannot Process!")
        exit()

    roll=[]
    for items in response_list:
        roll.append(items[6].upper())
        pos=0
        neg=0
        no_attempt=0
        for options, answers in zip(items[7:], ANSWER[7:]):
            if options==answers:
                pos=pos+1
            elif options=='':
                no_attempt=no_attempt+1
            else:
                neg=neg+1
        positive_marks=pos*positive
        negative_marks=neg*negative
        marks=positive_marks+negative_marks
        items.append([pos, neg, no_attempt, positive_marks, negative_marks, max, str(marks)+'/'+str(max*positive)])

    absent=[]

    # data_1=pd.read_csv(csv_file_1)     
    # df_1=pd.DataFrame(data_1)
    #st.success("Saved File")
    with open(os.path.join("sample_input",csv_file_1.name),"r") as f: 
        reader=csv.reader(f)
        master_roll_list=list(reader)
    #master_roll_list=[]
    # for i in df_1.values:
    #     j=i.tolist()
    #     master_roll_list.append(j)
    master_roll_list.pop(0)
    for i in master_roll_list:
        if i[0] not in roll:
            absent.append(i)
        if i[0]=='ANSWER':
            absent.remove(i)

    st.text("")
    st.text("")

    if st.button('Generate Roll No. wise Marksheets'):
        if os.listdir('sample_output\\marksheet'):
            for files in os.listdir('sample_output\\marksheet'):
                os.remove('sample_output\\marksheet'+'\\'+files) 
        result = sheets(response_list, ANSWER, positive, negative, absent)
        st.write('Marksheets Generated!')

    if st.button('Generate Concise Marksheet'):
        result = concise(response_list, absent)
        st.write('Concise Marksheet Generated!')

    if st.button('Send E-Mail'):
        if len(os.listdir('sample_output\\marksheet'))>1:
            result = send_email(response_list)
            st.write('E-Mail Sent Successfully!')
        else:
            st.write('Please Generate Marksheets First')
                        
else:
    #st.write("Please upload .csv file:(")

    if st.button('Generate Roll No. wise Marksheets'):
        st.write('Please upload the required files!')

    if st.button('Generate Concise Marksheet'):
        st.write("Please upload the required files!")

    if st.button('Send E-Mail'):
        st.write('Please upload the required files!')