'''
We will be performing all the calculations first and then insert the values in Excel Workbook
'''

'''
PART A: Calculations
'''

import csv
import os
os.system('cls')

# Obtaining a list of roll numbers and names from names-roll.csv file
roll_list=[]
names=[]
with open("names-roll.csv", "r") as f:
    reader=csv.reader(f)
    for row in reader:
        roll_list.append(row[0])
        names.append(row[1])
roll_list.pop(0)
names.pop(0)

# We obtain a master list comprising information of each roll no. in a list as an element (overall_list)
overall_list=[]
i=0
with open("grades.csv", "r") as csvfile:
    grade_r=csv.reader(csvfile)
    grade_ori_list=list(grade_r) #list of rows in grade.csv
    grade_ori_list.pop(0)
    grade_list=grade_ori_list.copy()

    with open("subjects_master.csv", "r") as file:
        sub_r=csv.reader(file)
        sub_list=list(sub_r)
        sub_list.pop(0)
        for row in grade_list:
            row.append(row[4])
            row.pop(4)
            for lists in sub_list:
                if row[2] in lists:
                    row.insert(3, lists[1])
                    row.insert(4, lists[2])

    for item in roll_list:
        roll_no=[]
        while i<len(grade_list):
            if i>=len(grade_list):  break
            if grade_list[i][0]==item:
                roll_no.append(grade_list[i])
                i=i+1
            else:
                break            
        overall_list.append(roll_no)

# CPI and SPI calculation
grade_dict={'AA':10, 'AB':9, 'BB':8, ' BB':8, 'BC':7, 'CC':6, 'CD':5, 'DD':4, 'F':0, 'I':0, 'F*':0, 'DD*':4}

SPI=[]
CPI=[]
SPI_creds, CPI_creds=[], []
sems=[]

for items in overall_list:
    SPI_roll=[items[0][0]]
    CPI_roll=[items[0][0]]
    CPI_num, CPI_dem=0, 0
    C_den, S_den=[items[0][0]], [items[0][0]]
    sems_i=[items[0][0]]

    for i in range(1, 11):
        SPI_num_i, den_i=0, 0
        for info in items:
            if int(info[1])==i:
                SPI_num_i=SPI_num_i+(int(info[5])*grade_dict[info[7]])
                den_i=den_i+int(info[5])
                if str(i) not in sems_i:
                    sems_i.append(str(i))
        try:        
            SPI_i=SPI_num_i/den_i
            SPI_roundoff=round(SPI_i, 2)
            SPI_roll.append(SPI_roundoff)
            CPI_num=CPI_num+(SPI_i*den_i)
            CPI_dem=CPI_dem+den_i
            CPI_roundoff=round(CPI_num/CPI_dem, 2)
            CPI_roll.append(CPI_roundoff)
            S_den.append(den_i)
            C_den.append(CPI_dem)
        except:
            pass
         
    SPI.append(SPI_roll)
    CPI.append(CPI_roll)
    SPI_creds.append(S_den)
    CPI_creds.append(C_den)
    sems.append(sems_i)
    
'''
PART B: Adding information to Workbook
'''

from openpyxl import Workbook

header= ['Sl No.', 'Subject No.', 'Subject Name', 'L-T-P', 'Credit', 'Subject Type', 'Grade']

for j, roll, item, length, sem in zip(range(len(roll_list)), roll_list, overall_list, SPI, sems):
    wb=Workbook()

    # Sheet 1: Overall
    Overall=wb.active
    Overall.title='Overall'

    # Column 1
    Overall['A1']='Roll No.'
    Overall['A2']='Name of Student'
    Overall['A3']='Discipline'
    Overall['A4']='Semester No.'
    Overall['A5']='Semester wise Credit Taken'
    Overall['A6']='SPI'
    Overall['A7']='Total Credits Taken'
    Overall['A8']='CPI'

    # Other columns
    Overall['B1']=roll
    Overall['B2']=names[j]
    Overall['B3']=roll[4:6]

    for i in range(10):
        try:
            Overall.cell(row= 5, column= i+2, value= SPI_creds[j][i+1])
            Overall.cell(row= 4, column= i+2, value= int(sem[i+1]))
            Overall.cell(row= 6, column= i+2, value= SPI[j][i+1])
            Overall.cell(row= 7, column= i+2, value= CPI_creds[j][i+1])
            Overall.cell(row= 8, column= i+2, value= CPI[j][i+1])
        except:
            pass

    # Other Sheets
    for k in range(len(length)-1):
        if roll+".xlsx" not in wb:
            sheet=wb.create_sheet('Sheet')
            sheet.title='Sem'+sem[k+1]
            # Row 1 containing the header
            for heads, t in zip(header, range(len(header))):
                sheet.cell(row= 1, column= t+1, value= heads)

            # Rows below containg names and figures
            i=1
            for info in item:
                if info[1]==sem[k+1]:
                    i=i+1
                    sheet.cell(row=i, column=1, value=i-1)
                    sheet.cell(row=i, column=2, value=info[2])
                    sheet.cell(row=i, column=3, value=info[3])
                    sheet.cell(row=i, column=4, value=info[4])
                    sheet.cell(row=i, column=5, value=int(info[5]))
                    sheet.cell(row=i, column=6, value=info[6])
                    sheet.cell(row=i, column=7, value=info[7])
                    
    wb.save("output\\" + roll + ".xlsx")